from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import TextIOWrapper
import csv
from pathlib import Path

from django import forms
from openpyxl import load_workbook


class CartAddForm(forms.Form):
    qty = forms.DecimalField(min_value=Decimal('0.001'), max_digits=14, decimal_places=3, label='Количество')
    replace = forms.BooleanField(required=False, initial=False, label='Заменить количество')


class CheckoutForm(forms.Form):
    comment = forms.CharField(required=False, widget=forms.Textarea(attrs={'rows': 3}), label='Комментарий к заказу')


class OrderUploadForm(forms.Form):
    file = forms.FileField(label='Файл заказа')
    mode = forms.ChoiceField(
        choices=[('append', 'Добавить к корзине'), ('replace', 'Полностью заменить корзину')],
        initial='replace',
        label='Как загрузить',
    )


@dataclass
class UploadedOrderRow:
    product_code: str
    qty: Decimal


HEADER_CODE_ALIASES = {'код', 'product_code', 'code', 'артикул', 'sku'}
HEADER_QTY_ALIASES = {'qty', 'quantity', 'количество', 'qty_order'}


class UploadedOrderParser:
    @staticmethod
    def parse(uploaded_file) -> list[UploadedOrderRow]:
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix in {'.xlsx', '.xlsm'}:
            return UploadedOrderParser._parse_xlsx(uploaded_file)
        if suffix == '.csv':
            return UploadedOrderParser._parse_csv(uploaded_file)
        raise forms.ValidationError('Поддерживаются только CSV, XLSX и XLSM.')

    @staticmethod
    def _normalize_headers(headers: list[str]) -> dict[str, int]:
        normalized = {}
        for idx, header in enumerate(headers):
            normalized[str(header).strip().lower()] = idx
        return normalized

    @staticmethod
    def _extract_indices(headers: list[str]) -> tuple[int, int]:
        normalized = UploadedOrderParser._normalize_headers(headers)
        code_idx = qty_idx = None
        for alias in HEADER_CODE_ALIASES:
            if alias in normalized:
                code_idx = normalized[alias]
                break
        for alias in HEADER_QTY_ALIASES:
            if alias in normalized:
                qty_idx = normalized[alias]
                break
        if code_idx is None or qty_idx is None:
            raise forms.ValidationError('В файле должны быть колонки Код и Количество.')
        return code_idx, qty_idx

    @staticmethod
    def _to_decimal(value) -> Decimal:
        try:
            return Decimal(str(value).replace(',', '.').strip())
        except Exception as exc:
            raise forms.ValidationError(f'Не удалось прочитать количество: {value!r}') from exc

    @staticmethod
    def _parse_xlsx(uploaded_file) -> list[UploadedOrderRow]:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise forms.ValidationError('Файл пустой.')
        headers = [str(x or '').strip() for x in rows[0]]
        code_idx, qty_idx = UploadedOrderParser._extract_indices(headers)
        parsed: list[UploadedOrderRow] = []
        for row in rows[1:]:
            if row is None:
                continue
            code = str(row[code_idx] or '').strip()
            qty_raw = row[qty_idx]
            if not code or qty_raw in (None, ''):
                continue
            parsed.append(UploadedOrderRow(product_code=code, qty=UploadedOrderParser._to_decimal(qty_raw)))
        if not parsed:
            raise forms.ValidationError('В файле не найдено ни одной строки заказа.')
        return parsed

    @staticmethod
    def _parse_csv(uploaded_file) -> list[UploadedOrderRow]:
        wrapper = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig')
        reader = csv.reader(wrapper)
        rows = list(reader)
        if not rows:
            raise forms.ValidationError('Файл пустой.')
        headers = [str(x or '').strip() for x in rows[0]]
        code_idx, qty_idx = UploadedOrderParser._extract_indices(headers)
        parsed: list[UploadedOrderRow] = []
        for row in rows[1:]:
            if not row:
                continue
            code = str(row[code_idx] or '').strip() if code_idx < len(row) else ''
            qty_raw = row[qty_idx] if qty_idx < len(row) else ''
            if not code or qty_raw in (None, ''):
                continue
            parsed.append(UploadedOrderRow(product_code=code, qty=UploadedOrderParser._to_decimal(qty_raw)))
        if not parsed:
            raise forms.ValidationError('В CSV не найдено ни одной строки заказа.')
        return parsed
